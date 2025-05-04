from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook


INPUT_FILE = "../assets/AGRIBALYSE.xlsx"
OUTPUT_FILE = "../assets/AGRIBALYSE_v2.xlsx"


# Charger le fichier Excel
wb = load_workbook(INPUT_FILE)

print("WB created")

# Sélectionner la feuille spécifique par son nom
sheet_name = 'Synthese'
ws = wb[sheet_name]

print("WS created")

# Fonction pour vérifier si le mot 'erreur' apparaît dans une cellule
def contains_error(cell):
    if cell.value and isinstance(cell.value, str):
        return 'erreur' in cell.value.lower()  # Insensible à la casse
    return False


# Créer un nouveau classeur
new_wb = Workbook()

# Ajouter la feuille modifiée dans le nouveau classeur
new_ws = new_wb.active
new_ws.title = ws.title  # Copier le nom de la feuille


# Parcourir les lignes et ajouter celles qui ne contiennent pas le mot 'erreur'
for row in ws.iter_rows():
    # Vérifier si une cellule de la ligne contient le mot 'erreur'
    if not any(contains_error(cell) for cell in row):
        # Ajouter la ligne entière à la nouvelle feuille si 'erreur' n'est pas trouvé
        new_ws.append([cell.value for cell in row])

# Sauvegarder ce nouveau fichier Excel avec uniquement la feuille modifiée
new_wb.save(OUTPUT_FILE)

